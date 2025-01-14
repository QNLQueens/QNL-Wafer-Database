import tkinter as tk
from tkinter import simpledialog, messagebox, filedialog, ttk
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.pyplot as plt
import matplotlib.lines as lines
import matplotlib.patches as patches
import math
import os
import openpyxl
from database import *

class WaferApp:
    def __init__(self, parent):
        self.parent = parent
        self.frame = tk.Frame(self.parent)
        self.frame.grid(row=0, column=1, padx=10, pady=10)

        self.figure, self.ax = plt.subplots()
        self.canvas = FigureCanvasTkAgg(self.figure, master=self.frame)
        self.canvas.get_tk_widget().pack(fill=tk.BOTH, expand=1)

        self.ax.set_aspect('equal')
        self.ax.set_xlim(0, 1)
        self.ax.set_ylim(0, 1)
        self.ax.set_xticks([])
        self.ax.set_yticks([])

        self.canvas.mpl_connect("button_press_event", self.on_click)
        self.canvas.mpl_connect("motion_notify_event", self.on_motion)
        self.canvas.mpl_connect("button_release_event", self.on_release)

        self.lines = []
        self.polygons = []
        self.current_line = None
        self.start_x, self.start_y = None, None
        self.start_shape_x, self.start_shape_y = None, None
        self.shape_count = 0
        self.colcount = 0

        self.global_filepath = "all_wafers.xlsx"
        self.wafer_id = None
        self.year = None

        self.wafers_data = self.load_wafers_data()
        self.create_controls()

        self.year_combobox = None
        self.wafer_combobox = None
        self.create_comboboxes()

        self.draw_wafer()

    def load_wafers_data(self):
        """Loads wafer data from 'wafers.xlsx'."""
        con = load_most_recent()
        wafers_data = con.table('wafers').execute().to_dict()
        # wafers_data = {}
        # wafers_filepath = "wafers.xlsx"

        # if os.path.exists(wafers_filepath):
        #     workbook = openpyxl.load_workbook(wafers_filepath)
        #     sheet = workbook.active

        #     for row in sheet.iter_rows(min_row=2, values_only=True):
        #         year = row[0]  # The first column is the year
        #         wafer_id = row[1]  # The second column is the wafer ID
        #         if year and wafer_id:  # Ensure both values are present
        #             if year not in wafers_data:
        #                 wafers_data[year] = []
        #             wafers_data[year].append(wafer_id)
        return wafers_data


    def create_comboboxes(self):
        """Creates and places the year and wafer comboboxes."""
        control_frame = tk.Frame(self.frame)
        control_frame.pack(fill=tk.X)

        year_label = tk.Label(control_frame, text="Year:")
        year_label.pack(side=tk.LEFT, padx=5, pady=5)

        self.year_combobox = ttk.Combobox(control_frame, state="readonly")
        self.year_combobox.pack(side=tk.LEFT, padx=5, pady=5)
        self.year_combobox["values"] = list(self.wafers_data.keys())
        self.year_combobox.bind("<<ComboboxSelected>>", self.on_year_selected)

        wafer_label = tk.Label(control_frame, text="Wafer ID:")
        wafer_label.pack(side=tk.LEFT, padx=5, pady=5)

        self.wafer_combobox = ttk.Combobox(control_frame, state="readonly")
        self.wafer_combobox.pack(side=tk.LEFT, padx=5, pady=5)
        self.wafer_combobox.bind("<<ComboboxSelected>>", self.on_wafer_selected)

    def on_year_selected(self, event):
        """Populates the wafer combobox based on the selected year."""
        selected_year = self.year_combobox.get()
        self.year = selected_year
        self.wafer_combobox["values"] = self.wafers_data.get(selected_year, [])
        self.wafer_combobox.set("")  # Clear the wafer selection

    def on_wafer_selected(self, event):
        """Sets the wafer ID based on the selected wafer in the combobox."""
        selected_wafer = self.wafer_combobox.get()
        if selected_wafer:
            self.wafer_id = selected_wafer
            self.load_saved_shapes()

    def create_controls(self):
        control_frame = tk.Frame(self.frame)
        control_frame.pack(fill=tk.X)

        clear_button = tk.Button(control_frame, text="Clear", command=self.clear_lines)
        clear_button.pack(side=tk.LEFT, padx=5, pady=5)

        reset_button = tk.Button(control_frame, text="Reset", command=self.reset_wafer)
        reset_button.pack(side=tk.LEFT, padx=5, pady=5)

        save_button = tk.Button(control_frame, text="Save", command=self.save_html)
        save_button.pack(side=tk.LEFT, padx=5, pady=5)

    def draw_wafer(self):
        self.ax.clear()
        wafer_circle = plt.Circle((0.5, 0.5), 0.5, color='lightgrey', fill=True)
        self.ax.add_artist(wafer_circle)
        for line in self.lines:
            line_obj = lines.Line2D([line['x1'], line['x2']], [line['y1'], line['y2']], color='blue')
            self.ax.add_line(line_obj)
        for polygon in self.polygons:
            self.ax.add_patch(polygon)
        self.canvas.draw()

    def on_click(self, event):
        if event.inaxes != self.ax or not self.wafer_id:
            return
        if ((event.xdata-0.5)**2 + (event.ydata-0.5)**2) > 0.25:
            return  # Outside the wafer
        if not self.lines:
            self.start_shape_x, self.start_shape_y = event.xdata, event.ydata
        self.start_x, self.start_y = event.xdata, event.ydata
        self.current_line = lines.Line2D([self.start_x, self.start_x], [self.start_y, self.start_y], color='red')
        self.ax.add_line(self.current_line)
        self.canvas.draw()

    def on_motion(self, event):
        if self.current_line is None:
            return
        if event.inaxes != self.ax:
            return
        self.current_line.set_data([self.start_x, event.xdata], [self.start_y, event.ydata])
        self.canvas.draw()

    def on_release(self, event):
        if self.current_line is None or not self.wafer_id:
            return
        if event.inaxes != self.ax:
            return
        self.lines.append({
            'x1': self.start_x,
            'x2': event.xdata,
            'y1': self.start_y,
            'y2': event.ydata
        })
        self.current_line.set_color('blue')
        self.current_line = None

        if self.is_close_to_start(event.xdata, event.ydata):
            self.fill_shape()
            if len(self.lines) == 4:
                self.get_chip_data_and_save()
            self.lines = []

        self.start_x, self.start_y = None, None
        self.draw_wafer()

    def is_close_to_start(self, x, y, tolerance=0.01):
        if self.start_shape_x is None or self.start_shape_y is None:
            return False
        return math.hypot(x - self.start_shape_x, y - self.start_shape_y) < tolerance

    def fill_shape(self):
        colours = ["red", "orange", "yellow", "green", "blue", "purple"]
        vertices = [(line['x1'], line['y1']) for line in self.lines] + [(self.lines[-1]['x2'], self.lines[-1]['y2'])]
        if self.colcount == 6:
            self.colcount = 0
        polygon = patches.Polygon(vertices, closed=True, fill=True, color=colours[self.colcount], alpha=0.5)
        self.polygons.append(polygon)
        self.colcount += 1
        self.shape_count += 1
        self.ax.add_patch(polygon)
        self.canvas.draw()

    def get_chip_data_and_save(self):
        chipID = simpledialog.askstring("Input", "Enter Chip ID:", parent=self.parent)
        owner = simpledialog.askstring("Input", "Enter Owner:", parent=self.parent)
        device = simpledialog.askstring("Input", "Enter Device:", parent=self.parent)

        if not chipID or not owner or not device:
            messagebox.showerror("Error", "All fields must be filled out!")
            return

        x1, y1 = self.lines[0]['x2'], self.lines[0]['y2']
        x2, y2 = self.lines[1]['x2'], self.lines[1]['y2']
        x3, y3 = self.lines[2]['x2'], self.lines[2]['y2']
        x4, y4 = self.lines[3]['x2'], self.lines[3]['y2']

        if not os.path.exists(self.global_filepath):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            heading = ["Year", "Wafer ID", "Chip ID", "Owner", "Device", "x1", "y1", "x2", "y2", "x3", "y3", "x4", "y4"]
            sheet.append(heading)
            workbook.save(self.global_filepath)

        workbook = openpyxl.load_workbook(self.global_filepath)
        sheet = workbook.active
        sheet.append([self.year, self.wafer_id, chipID, owner, device, x1, y1, x2, y2, x3, y3, x4, y4])
        workbook.save(self.global_filepath)

        messagebox.showinfo("Saved", f"Data for Chip ID {chipID} saved successfully!")

    def load_saved_shapes(self):
        if not os.path.exists(self.global_filepath) or not self.wafer_id:
            return

        workbook = openpyxl.load_workbook(self.global_filepath)
        sheet = workbook.active

        self.polygons.clear()

        for row in sheet.iter_rows(min_row=2, values_only=True):
            year, wafer_id, chipID, owner, device, x1, y1, x2, y2, x3, y3, x4, y4 = row
            if year == self.year and wafer_id == self.wafer_id:
                vertices = [(x1, y1), (x2, y2), (x3, y3), (x4, y4)]
                polygon = patches.Polygon(vertices, closed=True, fill=True, color='blue', alpha=0.5)
                self.polygons.append(polygon)
                self.ax.add_patch(polygon)
                self.shape_count += 1

        self.canvas.draw()

    def clear_lines(self):
        self.lines = []
        self.polygons = []
        self.ax.clear()
        self.draw_wafer()

    def reset_wafer(self):
        confirm = messagebox.askyesno("Reset", "Are you sure you want to reset the wafer? All unsaved data will be lost.")
        if confirm:
            self.clear_lines()
            self.shape_count = 0

    def save_html(self):
        messagebox.showinfo("Save", "Save functionality to be implemented.")

if __name__ == "__main__":
    root = tk.Tk()
    root.title("Wafer Drawing App")
    app = WaferApp(root)
    root.mainloop()
